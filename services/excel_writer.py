from pathlib import Path
from openpyxl import load_workbook


class ExcelWriter:

    def __init__(self):

        self.archivo = (
            Path(__file__).resolve().parent.parent
            / "data"
            / "condiciones.xlsx"
        )

    # ==========================================================
    # GUARDAR NUEVO REPORTE
    # ==========================================================

    def guardar_reporte(

        self,

        numero,

        fecha_reporte,

        accion_condicion,

        detalle,

        riesgos,

        priorizacion,

        area,

        lugar_maquina_equipo,

        responsable_area,

        accion_correctiva,

        responsable_cierre,

        fecha_propuesta_cierre,

        fecha_real_cierre,

        estado,

        evidencia_fotografica=""

    ):

        # ==========================================================
        # ABRIR EXCEL
        # ==========================================================

        wb = load_workbook(self.archivo)

        # Usar específicamente la hoja Base
        ws = wb["Base"]

        # Siguiente fila disponible
        fila = ws.max_row + 1

        # ==========================================================
        # GENERAR NÚMERO AUTOMÁTICO
        # ==========================================================

        numero = fila - 2

        # ==========================================================
        # GUARDAR DATOS
        # ==========================================================

        ws.cell(fila, 1).value = numero

        ws.cell(fila, 2).value = fecha_reporte

        ws.cell(fila, 3).value = accion_condicion

        ws.cell(fila, 4).value = detalle

        ws.cell(fila, 5).value = riesgos

        ws.cell(fila, 6).value = priorizacion

        ws.cell(fila, 7).value = area

        ws.cell(fila, 8).value = lugar_maquina_equipo

        ws.cell(fila, 9).value = responsable_area

        ws.cell(fila, 10).value = accion_correctiva

        ws.cell(fila, 11).value = responsable_cierre

        ws.cell(fila, 12).value = fecha_propuesta_cierre

        ws.cell(fila, 13).value = fecha_real_cierre

        ws.cell(fila, 14).value = estado

        ws.cell(fila, 15).value = evidencia_fotografica

        # ==========================================================
        # GUARDAR Y CERRAR
        # ==========================================================

        wb.save(self.archivo)

        wb.close()
